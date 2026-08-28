"""Tests del modulo de reportes y exportacion Excel."""

from datetime import date
from decimal import Decimal

from django.test import TestCase
from django.urls import reverse

from categorias.models import Categoria
from core.models import ConfiguracionSistema, TipoMovimiento
from movimientos.models import Movimiento
from openpyxl import load_workbook

from reportes.services import generar_reporte
from usuarios.models import Usuario


class ReportesServicesTests(TestCase):
    def setUp(self):
        ConfiguracionSistema.objects.create(saldo_inicial=Decimal("1000.00"))
        cat_ing = Categoria.objects.create(nombre="Ventas", tipo=TipoMovimiento.INGRESO)
        cat_egr = Categoria.objects.create(nombre="Compras", tipo=TipoMovimiento.EGRESO)
        # Antes del rango
        Movimiento.objects.create(
            tipo=TipoMovimiento.INGRESO, valor=Decimal("500.00"),
            fecha=date(2026, 7, 1), concepto="previo", categoria=cat_ing,
        )
        # Dentro del rango
        Movimiento.objects.create(
            tipo=TipoMovimiento.INGRESO, valor=Decimal("2000.00"),
            fecha=date(2026, 8, 10), concepto="Agosto 1", categoria=cat_ing,
        )
        Movimiento.objects.create(
            tipo=TipoMovimiento.EGRESO, valor=Decimal("800.00"),
            fecha=date(2026, 8, 15), concepto="Agosto egreso", categoria=cat_egr,
        )

    def test_genera_reporte_basico(self):
        r = generar_reporte(date(2026, 8, 1), date(2026, 8, 31))
        self.assertEqual(r["ingresos"], Decimal("2000.00"))
        self.assertEqual(r["egresos"], Decimal("800.00"))
        self.assertEqual(r["balance"], Decimal("1200.00"))
        # Saldo inicial periodo = 1000 (global) + 500 (previo) = 1500
        self.assertEqual(r["saldo_inicial_periodo"], Decimal("1500.00"))
        self.assertEqual(r["saldo_final"], Decimal("2700.00"))

    def test_ingresos_por_categoria(self):
        r = generar_reporte(date(2026, 8, 1), date(2026, 8, 31))
        self.assertEqual(len(r["ingresos_por_categoria"]), 1)
        self.assertEqual(r["ingresos_por_categoria"][0]["categoria__nombre"], "Ventas")
        self.assertEqual(r["ingresos_por_categoria"][0]["total"], Decimal("2000.00"))


class ReportesViewsTests(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(username="admin", password="alsi2026")
        self.client.login(username="admin", password="alsi2026")

    def test_reporte_view_carga(self):
        resp = self.client.get(reverse("reportes:reporte"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Reporte financiero")

    def test_exportar_movimientos_xlsx(self):
        cat = Categoria.objects.create(nombre="X", tipo=TipoMovimiento.INGRESO)
        Movimiento.objects.create(
            tipo=TipoMovimiento.INGRESO, valor=Decimal("100.00"),
            fecha=date(2026, 8, 20), concepto="Venta test", categoria=cat,
        )
        resp = self.client.get(reverse("reportes:exportar_movimientos"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # Verificar contenido
        wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[4]]
        self.assertIn("Fecha", headers)
        self.assertIn("Tipo", headers)
        self.assertIn("Valor", headers)

    def test_exportar_reporte_xlsx_tiene_hojas(self):
        resp = self.client.get(reverse("reportes:exportar_reporte"))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
        nombres = wb.sheetnames
        self.assertIn("Reporte financiero", nombres)
        self.assertIn("Ingresos por categoria", nombres)
        self.assertIn("Egresos por categoria", nombres)
        self.assertIn("Detalle movimientos", nombres)

    def test_exportar_ingresos_solo_ingresos(self):
        cat = Categoria.objects.create(nombre="X", tipo=TipoMovimiento.INGRESO)
        Movimiento.objects.create(
            tipo=TipoMovimiento.INGRESO, valor=Decimal("100"),
            fecha=date.today(), concepto="i", categoria=cat,
        )
        cat_e = Categoria.objects.create(nombre="Y", tipo=TipoMovimiento.EGRESO)
        Movimiento.objects.create(
            tipo=TipoMovimiento.EGRESO, valor=Decimal("50"),
            fecha=date.today(), concepto="e", categoria=cat_e,
        )
        resp = self.client.get(reverse("reportes:exportar_ingresos"))
        wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
        ws = wb.active
        tipos = [ws.cell(row=r, column=2).value for r in range(5, ws.max_row + 1)]
        self.assertTrue(all(t == "Ingreso" for t in tipos if t))
