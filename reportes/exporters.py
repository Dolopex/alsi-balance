"""Exportadores a Excel (.xlsx) usando openpyxl."""

from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import TipoMovimiento


# ---- Estilos comunes --------------------------------------------------------

COLOR_HEADER_BG = "0B3666"
COLOR_HEADER_FG = "FFFFFF"
COLOR_TITLE = "0B3666"
COLOR_BAND_LIGHT = "F8FAFC"
COLOR_BAND_ALT = "FFFFFF"
COLOR_TOTAL_BG = "EAF1FB"
COLOR_TOTAL_FG = "0B3666"
COLOR_BORDER = "CBD5E1"
COLOR_INCOME = "065F46"
COLOR_EXPENSE = "991B1B"
COLOR_INFO = "1E40AF"

HEADER_FILL = PatternFill("solid", fgColor=COLOR_HEADER_BG)
HEADER_FONT = Font(bold=True, color=COLOR_HEADER_FG, size=11)
TITLE_FONT = Font(bold=True, size=18, color=COLOR_TITLE)
SUBTITLE_FONT = Font(italic=True, size=10, color="64748B")
META_FONT = Font(size=10, color="475569")
BAND_LIGHT_FILL = PatternFill("solid", fgColor=COLOR_BAND_LIGHT)
TOTAL_FILL = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
TOTAL_FONT = Font(bold=True, color=COLOR_TOTAL_FG, size=11)
INCOME_FONT = Font(bold=True, color=COLOR_INCOME)
EXPENSE_FONT = Font(bold=True, color=COLOR_EXPENSE)

THIN = Side(border_style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FORMAT = '"$"#,##0.00;[Red]-"$"#,##0.00'
MONEY_FORMAT_TOTAL = '"$"#,##0.00'
DATE_FORMAT = "dd/mm/yyyy"
DATETIME_FORMAT = "dd/mm/yyyy hh:mm:ss"


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 50):
    """Ajusta el ancho de las columnas al contenido."""
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except AttributeError:
            continue
        length = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = max(min_width, min(length + 2, max_width))


def _estilo_header(ws: Worksheet, row: int, ncols: int):
    """Aplica estilo al header de una tabla."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[row].height = 24


def _aplicar_bandas(ws: Worksheet, primera_fila: int, ultima_fila: int, ncols: int):
    """Alterna colores de fondo en las filas para mejor legibilidad."""
    for fila in range(primera_fila, ultima_fila + 1):
        if (fila - primera_fila) % 2 == 1:
            for col in range(1, ncols + 1):
                cell = ws.cell(row=fila, column=col)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BAND_LIGHT_FILL
        for col in range(1, ncols + 1):
            ws.cell(row=fila, column=col).border = BORDER


def _header_empresa(ws: Worksheet, titulo: str, subtitulo: str = ""):
    """Escribe el header corporativo en las primeras filas."""
    ws["A1"] = "ALSI BALANCE"
    ws["A1"].font = Font(bold=True, size=20, color=COLOR_TITLE)
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Agropesquera La Sinuana S.A.S."
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 16

    ws["A3"] = titulo
    ws["A3"].font = Font(bold=True, size=14, color=COLOR_TITLE)
    ws.row_dimensions[3].height = 22

    if subtitulo:
        ws["A4"] = subtitulo
        ws["A4"].font = META_FONT
        ws.row_dimensions[4].height = 18

    ws["A5"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A5"].font = META_FONT


def _formato_moneda(cell):
    cell.number_format = MONEY_FORMAT
    cell.alignment = Alignment(horizontal="right")


def _formato_fecha(cell):
    cell.number_format = DATE_FORMAT
    cell.alignment = Alignment(horizontal="center")


# ---- Exportadores -----------------------------------------------------------


def exportar_movimientos(qs, titulo: str = "Movimientos", subtitulo: str = "") -> Workbook:
    """Exporta una queryset de movimientos a un Excel con formato profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    fila_header = 6 if subtitulo else 5
    _header_empresa(ws, titulo, subtitulo)

    headers = [
        "Fecha", "Hora", "Tipo", "Concepto", "Categoria",
        "Cuenta origen", "Cuenta destino", "Destinatario",
        "Valor", "Origen", "Estado", "Referencia",
    ]
    ws.append([])
    ws.append(headers)
    _estilo_header(ws, ws.max_row, len(headers))

    primera_fila = ws.max_row + 1
    totales_por_tipo = {TipoMovimiento.INGRESO: Decimal("0"), TipoMovimiento.EGRESO: Decimal("0")}

    for m in qs:
        fila_datos = [
            m.fecha,
            m.hora,
            m.get_tipo_display(),
            m.concepto or "",
            m.categoria.nombre if m.categoria else "",
            m.cuenta or "",
            m.cuenta_destino or "",
            m.nombre_destinatario or m.tercero or "",
            float(m.valor),
            m.get_origen_display(),
            m.get_estado_conciliacion_display(),
            m.referencia or "",
        ]
        ws.append(fila_datos)
        last_row = ws.max_row
        # Formatos especificos
        _formato_fecha(ws.cell(row=last_row, column=1))
        if m.hora:
            ws.cell(row=last_row, column=2).number_format = "hh:mm:ss"
            ws.cell(row=last_row, column=2).alignment = Alignment(horizontal="center")
        _formato_moneda(ws.cell(row=last_row, column=9))

        # Color del tipo
        tipo_cell = ws.cell(row=last_row, column=3)
        if m.tipo == TipoMovimiento.INGRESO:
            tipo_cell.font = INCOME_FONT
        elif m.tipo == TipoMovimiento.EGRESO:
            tipo_cell.font = EXPENSE_FONT

        # Acumular totales
        totales_por_tipo[m.tipo] = totales_por_tipo.get(m.tipo, Decimal("0")) + m.valor

    ultima_fila = ws.max_row
    _aplicar_bandas(ws, primera_fila, ultima_fila, len(headers))

    # Fila de totales
    ws.append([])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTALES")
    ws.cell(row=total_row, column=1).font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL

    ws.cell(row=total_row, column=3, value=f"Ingresos: ${totales_por_tipo[TipoMovimiento.INGRESO]:,.2f}")
    ws.cell(row=total_row, column=3).font = INCOME_FONT
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row, column=9, value=f"Egresos: ${totales_por_tipo[TipoMovimiento.EGRESO]:,.2f}")
    ws.cell(row=total_row, column=9).font = EXPENSE_FONT
    _formato_moneda(ws.cell(row=total_row, column=9))

    ws.cell(row=total_row, column=11, value=f"Balance: ${totales_por_tipo[TipoMovimiento.INGRESO] - totales_por_tipo[TipoMovimiento.EGRESO]:,.2f}")
    ws.cell(row=total_row, column=11).font = TOTAL_FONT
    _formato_moneda(ws.cell(row=total_row, column=11))

    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = BORDER

    _autosize(ws, min_width=12, max_width=40)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=primera_fila, column=1)

    # Filtros automaticos (Excel los ve como headers)
    ws.auto_filter.ref = f"A{primera_fila - 1}:{get_column_letter(len(headers))}{ultima_fila}"

    return wb


def exportar_ingresos(qs, subtitulo: str = "") -> Workbook:
    """Exporta solo ingresos de un periodo."""
    return exportar_movimientos(
        qs.filter(tipo=TipoMovimiento.INGRESO),
        "Reporte de Ingresos",
        subtitulo,
    )


def exportar_egresos(qs, subtitulo: str = "") -> Workbook:
    """Exporta solo egresos de un periodo."""
    return exportar_movimientos(
        qs.filter(tipo=TipoMovimiento.EGRESO),
        "Reporte de Egresos",
        subtitulo,
    )


def exportar_reporte_financiero(reporte: dict) -> Workbook:
    """Genera un reporte financiero completo en multiples hojas."""
    wb = Workbook()

    # ---- Hoja resumen ----
    ws = wb.active
    ws.title = "Resumen"

    _header_empresa(
        ws,
        "Reporte Financiero",
        f"Periodo: {reporte['fecha_desde']:%d/%m/%Y} a {reporte['fecha_hasta']:%d/%m/%Y}",
    )

    fila_actual = 6
    ws.cell(row=fila_actual, column=1, value="Concepto")
    ws.cell(row=fila_actual, column=2, value="Valor")
    _estilo_header(ws, fila_actual, 2)
    fila_actual += 1

    filas_resumen = [
        ("Saldo inicial del periodo", reporte["saldo_inicial_periodo"], False),
        ("(+) Ingresos", reporte["ingresos"], True),
        ("(-) Egresos", reporte["egresos"], True),
        ("(=) Balance del periodo", reporte["balance"], True),
        ("Saldo final", reporte["saldo_final"], True),
    ]

    for etiqueta, valor, es_total in filas_resumen:
        ws.cell(row=fila_actual, column=1, value=etiqueta)
        cell = ws.cell(row=fila_actual, column=2, value=float(valor))
        _formato_moneda(cell)
        if es_total:
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            ws.cell(row=fila_actual, column=1).font = TOTAL_FONT
        for col in (1, 2):
            ws.cell(row=fila_actual, column=col).border = BORDER
        fila_actual += 1

    _autosize(ws, min_width=35, max_width=50)

    # ---- Hoja ingresos por categoria ----
    ws2 = wb.create_sheet("Ingresos por categoria")
    _header_empresa(ws2, "Ingresos por Categoria", f"Periodo: {reporte['fecha_desde']:%d/%m/%Y} a {reporte['fecha_hasta']:%d/%m/%Y}")
    ws2.append([])
    ws2.append(["Categoria", "Total ingresos"])
    _estilo_header(ws2, ws2.max_row, 2)
    primera = ws2.max_row + 1
    for row in reporte["ingresos_por_categoria"]:
        ws2.append([row["categoria__nombre"] or "(sin categoria)", float(row["total"] or 0)])
        _formato_moneda(ws2.cell(row=ws2.max_row, column=2))
    _aplicar_bandas(ws2, primera, ws2.max_row, 2)
    _autosize(ws2, min_width=30, max_width=50)

    # ---- Hoja egresos por categoria ----
    ws3 = wb.create_sheet("Egresos por categoria")
    _header_empresa(ws3, "Egresos por Categoria", f"Periodo: {reporte['fecha_desde']:%d/%m/%Y} a {reporte['fecha_hasta']:%d/%m/%Y}")
    ws3.append([])
    ws3.append(["Categoria", "Total egresos"])
    _estilo_header(ws3, ws3.max_row, 2)
    primera = ws3.max_row + 1
    for row in reporte["egresos_por_categoria"]:
        ws3.append([row["categoria__nombre"] or "(sin categoria)", float(row["total"] or 0)])
        _formato_moneda(ws3.cell(row=ws3.max_row, column=2))
    _aplicar_bandas(ws3, primera, ws3.max_row, 2)
    _autosize(ws3, min_width=30, max_width=50)

    # ---- Hoja detalle de movimientos ----
    ws4 = wb.create_sheet("Detalle movimientos")
    headers = ["Fecha", "Hora", "Tipo", "Concepto", "Categoria", "Valor"]
    _header_empresa(ws4, "Detalle de Movimientos", f"Periodo: {reporte['fecha_desde']:%d/%m/%Y} a {reporte['fecha_hasta']:%d/%m/%Y}")
    ws4.append([])
    ws4.append(headers)
    _estilo_header(ws4, ws4.max_row, len(headers))
    primera = ws4.max_row + 1
    for m in reporte["movimientos"]:
        ws4.append([
            m.fecha,
            m.hora,
            m.get_tipo_display(),
            m.concepto or "",
            m.categoria.nombre if m.categoria else "",
            float(m.valor),
        ])
        last = ws4.max_row
        _formato_fecha(ws4.cell(row=last, column=1))
        if m.hora:
            ws4.cell(row=last, column=2).number_format = "hh:mm:ss"
        _formato_moneda(ws4.cell(row=last, column=6))
    _aplicar_bandas(ws4, primera, ws4.max_row, len(headers))
    _autosize(ws4)
    ws4.freeze_panes = ws4.cell(row=primera, column=1)

    return wb


def workbook_to_response(wb: Workbook, filename: str):
    """Helper para que las vistas serialicen el workbook a una respuesta HTTP."""
    from django.http import HttpResponse

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response